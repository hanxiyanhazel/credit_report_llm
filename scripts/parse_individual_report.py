#!/usr/bin/env python3
"""Parse personal credit-report XML into the table-oriented standard JSON schema."""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any
import xml.etree.ElementTree as ET


SCHEMA_VERSION = "0.2.0"
DICTIONARY_FILE = "征信平台信用报告表结构明细V0.3.xlsx"
TABLE_LABEL_SOURCE = DICTIONARY_FILE

TABLE_NAMES = {
    "PA01": "报告头信息单元",
    "PB01": "身份信息单元",
    "PB02": "配偶信息单元",
    "PB03": "居住信息单元",
    "PB04": "职业信息单元",
    "PC05": "查询记录概要信息单元",
    "PD01": "信贷账户信息单元",
    "PD03": "对外担保信息单元",
    "PH01": "查询记录明细单元",
    "PCE": "公共记录信息单元",
    "POS": "异议信息单元",
}

FALLBACK_FIELD_LABELS = {
    "PA01CH": "其他证件信息明细",
    "PB01BH": "手机号码信息明细",
    "PA01AI01": "报告编号",
    "PA01AR01": "报告时间",
    "PA01BQ01": "被查询者姓名",
    "PA01BI01": "被查询者证件号码",
    "PA01BD01": "被查询者证件类型",
    "PA01BI02": "查询机构代码",
    "PA01BD02": "查询原因代码",
    "PA01CS01": "身份识别个数",
    "PA01CD01": "证件类型",
    "PA01CI01": "证件号码",
    "PB01AD01": "性别",
    "PB01AR01": "出生日期",
    "PB01AD02": "学历",
    "PB01AD03": "学位",
    "PB01AD04": "就业状况",
    "PB01AD05": "国籍",
    "PB01AQ01": "电子邮箱",
    "PB01AQ02": "通讯地址",
    "PB01AQ03": "户籍地址",
    "PB01BS01": "手机号码个数",
    "PB01BQ01": "手机号码",
    "PB01BR01": "信息更新日期",
    "PB020D01": "婚姻状况",
    "PB020Q01": "配偶姓名",
    "PB020D02": "配偶证件类型",
    "PB020I01": "配偶证件号码",
    "PB020Q02": "配偶工作单位",
    "PB020Q03": "配偶联系电话",
    "PB030D01": "居住状况",
    "PB030Q01": "居住地址",
    "PB030Q02": "住宅电话",
    "PB030R01": "信息更新日期",
    "PB040D01": "单位性质补充类型",
    "PB040Q01": "工作单位",
    "PB040D02": "单位性质",
    "PB040D03": "行业",
    "PB040Q02": "单位地址",
    "PB040Q03": "单位电话",
    "PB040D04": "职业",
    "PB040D05": "职务",
    "PB040D06": "职称",
    "PB040R01": "进入本单位年份",
    "PB040R02": "信息更新日期",
    "PC05AR01": "最近一次查询日期",
    "PC05AD01": "最近一次查询机构类型",
    "PC05AI01": "最近一次查询机构代码",
    "PC05AQ01": "最近一次查询原因",
    "PC05BS01": "最近一个月信用卡审批查询次数",
    "PC05BS02": "最近一个月贷后管理查询次数",
    "PC05BS03": "最近一个月贷款审批查询次数",
    "PC05BS04": "最近一个月担保资格审查查询次数",
    "PC05BS05": "最近一个月特约商户实名审查查询次数",
    "PC05BS06": "最近两年查询总次数",
    "PC05BS07": "最近两年查询机构总数",
    "PC05BS08": "最近两年贷后管理查询次数",
    "PD01ZH": "特殊交易明细",
    "PD01ZD01": "特殊交易类型",
    "PD01ZQ01": "特殊交易说明",
    "PD01ZR01": "特殊交易日期",
    "PH010R01": "查询日期",
    "PH010D01": "查询机构类型",
    "PH010Q02": "查询机构代码",
    "PH010Q03": "查询原因",
    "PG010D01": "标注及声明大类",
    "PG010D02": "标注及声明细类",
    "PG010S01": "异议条数",
    "PG010D03": "状态分类",
    "PG010Q01": "标注或声明内容",
    "PG010R01": "添加日期",
}

CODED_FIELDS = {
    "PA01BD01",
    "PA01BD02",
    "PA01CD01",
    "PB01AD01",
    "PB01AD02",
    "PB01AD03",
    "PB01AD04",
    "PB01AD05",
    "PB020D01",
    "PB020D02",
    "PB030D01",
    "PB040D01",
    "PB040D02",
    "PB040D03",
    "PB040D04",
    "PB040D05",
    "PB040D06",
    "PC05AD01",
    "PC05AQ01",
    "PD01AD01",
    "PD01AD02",
    "PD01AD03",
    "PD01AD05",
    "PD01AD06",
    "PD01AD07",
    "PD01AD08",
    "PD01AD10",
    "PD01BD01",
    "PD01BD03",
    "PD01BD04",
    "PD01ZD01",
    "PD03AD01",
    "PD03AD02",
    "PD03AD05",
    "PD03AD07",
    "PH010D01",
    "PH010Q03",
    "PG010D01",
    "PG010D02",
    "PG010D03",
    "PF03AD01",
}

DATE_FIELDS = {
    "PA01AR01",
    "PB01AR01",
    "PB01BR01",
    "PB030R01",
    "PB040R02",
    "PC05AR01",
    "PD01AR01",
    "PD01AR02",
    "PD01BR01",
    "PD01BR02",
    "PD01BR03",
    "PD01ZR01",
    "PD03AR01",
    "PD03AR02",
    "PH010R01",
    "PG010R01",
    "PF03AR01",
    "PF03AR02",
}

NUMERIC_FIELDS = {
    "PA01CS01",
    "PB01BS01",
    "PC05BS01",
    "PC05BS02",
    "PC05BS03",
    "PC05BS04",
    "PC05BS05",
    "PC05BS06",
    "PC05BS07",
    "PC05BS08",
    "PD01AJ01",
    "PD01AJ02",
    "PD01AJ03",
    "PD01BJ01",
    "PD01BJ02",
    "PD01BJ03",
    "PD03AJ01",
    "PD03AJ02",
    "PD03AJ03",
    "PG010S01",
    "PF03AJ01",
    "PF03AJ02",
}


def clean(value: str | None) -> str | None:
    if value is None:
        return None
    value = value.strip()
    return value or None


def parse_date_like(value: str | None) -> date | None:
    value = clean(value)
    if value is None:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y"):
        try:
            parsed = datetime.strptime(value, fmt)
        except ValueError:
            continue
        if fmt == "%Y":
            return date(parsed.year, 1, 1)
        if fmt == "%Y-%m":
            return date(parsed.year, parsed.month, 1)
        return parsed.date()
    return None


def iso_date(value: str | None) -> str | None:
    parsed = parse_date_like(value)
    return parsed.isoformat() if parsed else clean(value)


def to_number(value: str | None) -> int | float | None:
    value = clean(value)
    if value is None:
        return None
    try:
        if "." in value:
            return float(value)
        return int(value)
    except ValueError:
        return None


def transform_value(field_code: str, value: str | None) -> Any:
    if field_code in DATE_FIELDS:
        return iso_date(value)
    if field_code in NUMERIC_FIELDS:
        return to_number(value)
    return clean(value)


def load_dictionary_metadata(base_dir: Path) -> tuple[dict[str, str], dict[str, str]]:
    field_labels = dict(FALLBACK_FIELD_LABELS)
    field_types: dict[str, str] = {}
    workbook_path = base_dir / "data_recource" / DICTIONARY_FILE
    try:
        from openpyxl import load_workbook
    except ImportError:
        return field_labels, field_types
    if not workbook_path.exists():
        return field_labels, field_types

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["个人信用报告表结构"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        _, _, field_code, field_label, field_type, *_ = row
        if not field_code:
            continue
        field_code = str(field_code).strip()
        if field_label:
            field_labels[field_code] = str(field_label).strip()
        if field_type:
            field_types[field_code] = str(field_type).strip()
    return field_labels, field_types


def make_trace(
    xml_path: Path,
    *,
    source_section: str,
    source_table: str,
    field_code: str | None = None,
    source_id: str | None = None,
) -> dict[str, Any]:
    fields = [field_code] if field_code else []
    return {
        "source_file": str(xml_path),
        "source_section": source_section,
        "source_table": source_table,
        "source_fields": fields,
        "source_id": source_id,
    }


def plain_field(
    field_code: str,
    field_name: str,
    value: Any,
    value_meta: dict[str, Any],
) -> dict[str, Any]:
    return {
        "field_code": field_code,
        "field_name": field_name,
        "value": value,
        "value_meta": value_meta,
    }


def coded_field(
    field_code: str,
    field_name: str,
    code: str | None,
    value_meta: dict[str, Any],
) -> dict[str, Any]:
    return {
        "field_code": field_code,
        "field_name": field_name,
        "code": clean(code),
        "label": None,
        "label_source": "unknown",
        "interpretation_status": "unresolved",
        "value_meta": value_meta,
    }


def field_from_text(
    field_code: str,
    raw_value: str | None,
    *,
    xml_path: Path,
    source_section: str,
    source_table: str,
    source_id: str | None,
    field_labels: dict[str, str],
) -> dict[str, Any]:
    field_name = field_labels.get(field_code, field_code)
    meta = {
        "value_source": "xml_direct",
        "trace": make_trace(
            xml_path,
            source_section=source_section,
            source_table=source_table,
            field_code=field_code,
            source_id=source_id,
        ),
        "extra_fields": {},
    }
    if field_code in CODED_FIELDS:
        return coded_field(field_code, field_name, raw_value, meta)
    return plain_field(field_code, field_name, transform_value(field_code, raw_value), meta)


def nested_record(record_index: int, fields: dict[str, Any], record_id: str | None = None) -> dict[str, Any]:
    return {
        "record_index": record_index,
        "record_id": record_id,
        "fields": fields,
    }


def logical_table_single(table_code: str, fields: dict[str, Any]) -> dict[str, Any]:
    return {
        "table_code": table_code,
        "table_name": TABLE_NAMES[table_code],
        "table_label_source": TABLE_LABEL_SOURCE,
        "fields": fields,
    }


def logical_table_repeated(
    table_code: str,
    records: list[dict[str, Any]],
    *,
    summary_fields: dict[str, Any] | None = None,
) -> dict[str, Any]:
    table = {
        "table_code": table_code,
        "table_name": TABLE_NAMES[table_code],
        "table_label_source": TABLE_LABEL_SOURCE,
        "records": records,
    }
    if summary_fields:
        table["summary_fields"] = summary_fields
    return table


def leaf_fields(
    node: ET.Element | None,
    *,
    xml_path: Path,
    source_section: str,
    source_table: str,
    source_id: str | None,
    field_labels: dict[str, str],
    skip_tags: set[str] | None = None,
) -> dict[str, Any]:
    fields: dict[str, Any] = {}
    if node is None:
        return fields
    for child in list(node):
        if len(child) != 0:
            continue
        if skip_tags and child.tag in skip_tags:
            continue
        fields[child.tag] = field_from_text(
            child.tag,
            child.text,
            xml_path=xml_path,
            source_section=source_section,
            source_table=source_table,
            source_id=source_id,
            field_labels=field_labels,
        )
    return fields


def repeated_value_field(
    field_code: str,
    field_name: str,
    records: list[dict[str, Any]],
    *,
    xml_path: Path,
    source_section: str,
    source_table: str,
    source_id: str | None,
) -> dict[str, Any]:
    return plain_field(
        field_code,
        field_name,
        records,
        {
            "value_source": "xml_direct",
            "trace": make_trace(
                xml_path,
                source_section=source_section,
                source_table=source_table,
                field_code=field_code,
                source_id=source_id,
            ),
            "extra_fields": {},
        },
    )


def build_pa01(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    pa01 = root.find(".//PA01")
    fields: dict[str, Any] = {}
    if pa01 is None:
        return logical_table_single("PA01", fields)

    fields.update(
        leaf_fields(
            pa01.find("PA01A"),
            xml_path=xml_path,
            source_section="PRH/PA01/PA01A",
            source_table="PCR_PA01A",
            source_id=None,
            field_labels=field_labels,
        )
    )
    fields.update(
        leaf_fields(
            pa01.find("PA01B"),
            xml_path=xml_path,
            source_section="PRH/PA01/PA01B",
            source_table="PCR_PA01A",
            source_id=None,
            field_labels=field_labels,
        )
    )
    pa01c = pa01.find("PA01C")
    if pa01c is not None:
        fields.update(
            leaf_fields(
                pa01c,
                xml_path=xml_path,
                source_section="PRH/PA01/PA01C",
                source_table="PCR_PA01A",
                source_id=None,
                field_labels=field_labels,
            )
        )
        nested: list[dict[str, Any]] = []
        for idx, node in enumerate(pa01c.findall("PA01CH"), start=1):
            record_fields = leaf_fields(
                node,
                xml_path=xml_path,
                source_section="PRH/PA01/PA01C/PA01CH",
                source_table="PCR_PA01A",
                source_id=f"PA01CH:{idx}",
                field_labels=field_labels,
            )
            nested.append(nested_record(idx, record_fields))
        if nested:
            fields["PA01CH"] = repeated_value_field(
                "PA01CH",
                field_labels.get("PA01CH", "其他证件信息明细"),
                nested,
                xml_path=xml_path,
                source_section="PRH/PA01/PA01C",
                source_table="PCR_PA01A",
                source_id=None,
            )
    return logical_table_single("PA01", fields)


def build_pb01(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    pb01 = root.find(".//PB01")
    fields: dict[str, Any] = {}
    if pb01 is None:
        return logical_table_single("PB01", fields)

    fields.update(
        leaf_fields(
            pb01.find("PB01A"),
            xml_path=xml_path,
            source_section="PIM/PB01/PB01A",
            source_table="PCR_PB01A",
            source_id=None,
            field_labels=field_labels,
        )
    )
    pb01b = pb01.find("PB01B")
    if pb01b is not None:
        fields.update(
            leaf_fields(
                pb01b,
                xml_path=xml_path,
                source_section="PIM/PB01/PB01B",
                source_table="PCR_PB01B",
                source_id=None,
                field_labels=field_labels,
                skip_tags={"PB01BH"},
            )
        )
        nested: list[dict[str, Any]] = []
        for idx, node in enumerate(pb01b.findall("PB01BH"), start=1):
            record_fields = leaf_fields(
                node,
                xml_path=xml_path,
                source_section="PIM/PB01/PB01B/PB01BH",
                source_table="PCR_PB01B",
                source_id=f"PB01BH:{idx}",
                field_labels=field_labels,
            )
            nested.append(nested_record(idx, record_fields))
        if nested:
            fields["PB01BH"] = repeated_value_field(
                "PB01BH",
                field_labels.get("PB01BH", "手机号码信息明细"),
                nested,
                xml_path=xml_path,
                source_section="PIM/PB01/PB01B",
                source_table="PCR_PB01B",
                source_id=None,
            )
    return logical_table_single("PB01", fields)


def build_pb02(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    for idx, node in enumerate(root.findall(".//PB02"), start=1):
        fields = leaf_fields(
            node,
            xml_path=xml_path,
            source_section="PIM/PB02",
            source_table="PCR_PB02",
            source_id=f"PB02:{idx}",
            field_labels=field_labels,
        )
        if fields:
            records.append(nested_record(idx, fields))
    return logical_table_repeated("PB02", records)


def build_pb03(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    for idx, node in enumerate(root.findall(".//PB03"), start=1):
        fields = leaf_fields(
            node,
            xml_path=xml_path,
            source_section="PRM/PB03",
            source_table="PCR_PB03",
            source_id=f"PB03:{idx}",
            field_labels=field_labels,
        )
        records.append(nested_record(idx, fields))
    return logical_table_repeated("PB03", records)


def build_pb04(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    for idx, node in enumerate(root.findall(".//PB04"), start=1):
        fields = leaf_fields(
            node,
            xml_path=xml_path,
            source_section="POM/PB04",
            source_table="PCR_PB04",
            source_id=f"PB04:{idx}",
            field_labels=field_labels,
        )
        records.append(nested_record(idx, fields, record_id=clean(node.findtext("PB040Q01"))))
    return logical_table_repeated("PB04", records)


def build_pc05(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    pc05 = root.find(".//PC05")
    fields: dict[str, Any] = {}
    if pc05 is None:
        return logical_table_single("PC05", fields)
    fields.update(
        leaf_fields(
            pc05.find("PC05A"),
            xml_path=xml_path,
            source_section="PQO/PC05/PC05A",
            source_table="PCR_PC05A",
            source_id=None,
            field_labels=field_labels,
        )
    )
    fields.update(
        leaf_fields(
            pc05.find("PC05B"),
            xml_path=xml_path,
            source_section="PQO/PC05/PC05B",
            source_table="PCR_PC05B",
            source_id=None,
            field_labels=field_labels,
        )
    )
    return logical_table_single("PC05", fields)


def build_pd01(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    for idx, pd01 in enumerate(root.findall(".//PD01"), start=1):
        fields: dict[str, Any] = {}
        fields.update(
            leaf_fields(
                pd01.find("PD01A"),
                xml_path=xml_path,
                source_section="PDA/PD01/PD01A",
                source_table="PCR_PD01A",
                source_id=f"PD01:{idx}",
                field_labels=field_labels,
            )
        )
        fields.update(
            leaf_fields(
                pd01.find("PD01B"),
                xml_path=xml_path,
                source_section="PDA/PD01/PD01B",
                source_table="PCR_PD01B",
                source_id=f"PD01:{idx}",
                field_labels=field_labels,
            )
        )
        pd01z = pd01.find("PD01Z")
        if pd01z is not None:
            nested: list[dict[str, Any]] = []
            for z_idx, node in enumerate(pd01z.findall("PD01ZH"), start=1):
                record_fields = leaf_fields(
                    node,
                    xml_path=xml_path,
                    source_section="PDA/PD01/PD01Z/PD01ZH",
                    source_table="PCR_PD01Z",
                    source_id=f"PD01:{idx}:PD01ZH:{z_idx}",
                    field_labels=field_labels,
                )
                nested.append(nested_record(z_idx, record_fields))
            if nested:
                fields["PD01ZH"] = repeated_value_field(
                    "PD01ZH",
                    field_labels.get("PD01ZH", "特殊交易明细"),
                    nested,
                    xml_path=xml_path,
                    source_section="PDA/PD01/PD01Z",
                    source_table="PCR_PD01Z",
                    source_id=f"PD01:{idx}",
                )
        record_id = None
        account_field = fields.get("PD01AI01")
        if account_field:
            record_id = clean(str(account_field.get("value")))
        records.append(nested_record(idx, fields, record_id=record_id))
    return logical_table_repeated("PD01", records)


def build_pd03(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    for idx, node in enumerate(root.findall(".//PD03"), start=1):
        fields = leaf_fields(
            node,
            xml_path=xml_path,
            source_section="PPO/PD03",
            source_table="PCR_PD03",
            source_id=f"PD03:{idx}",
            field_labels=field_labels,
        )
        records.append(nested_record(idx, fields))
    return logical_table_repeated("PD03", records)


def build_ph01(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    for idx, node in enumerate(root.findall(".//PH01"), start=1):
        fields = leaf_fields(
            node,
            xml_path=xml_path,
            source_section="POQ/PH01",
            source_table="PCR_PH01",
            source_id=f"PH01:{idx}",
            field_labels=field_labels,
        )
        records.append(nested_record(idx, fields))
    return logical_table_repeated("PH01", records)


def build_pce(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    for idx, node in enumerate(root.findall(".//PF03A"), start=1):
        fields = leaf_fields(
            node,
            xml_path=xml_path,
            source_section="PCE/PF03/PF03A",
            source_table="PCR_PF03A",
            source_id=f"PF03A:{idx}",
            field_labels=field_labels,
        )
        records.append(nested_record(idx, fields))
    return logical_table_repeated("PCE", records)


def build_pos(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    summary_fields: dict[str, Any] = {}
    for idx, node in enumerate(root.findall(".//PG01"), start=1):
        summary_fields.update(
            leaf_fields(
                node,
                xml_path=xml_path,
                source_section="POS/PG01",
                source_table="PCR_PG01",
                source_id=f"PG01:{idx}",
                field_labels=field_labels,
                skip_tags={"PG010H"},
            )
        )
        for h_idx, sub in enumerate(node.findall("PG010H"), start=1):
            fields = leaf_fields(
                sub,
                xml_path=xml_path,
                source_section="POS/PG01/PG010H",
                source_table="PCR_PG01H",
                source_id=f"PG01:{idx}:PG010H:{h_idx}",
                field_labels=field_labels,
            )
            records.append(nested_record(len(records) + 1, fields))
    return logical_table_repeated("POS", records, summary_fields=summary_fields or None)


def build_document(xml_path: Path) -> dict[str, Any]:
    base_dir = Path(__file__).resolve().parents[1]
    dictionary_path = base_dir / "data_recource" / DICTIONARY_FILE
    field_labels, _field_types = load_dictionary_metadata(base_dir)
    root = ET.parse(xml_path).getroot()

    document = {
        "schema_version": SCHEMA_VERSION,
        "report_type": "individual",
        "source_files": [
            {
                "path": str(xml_path),
                "role": "xml",
                "description": "个人征信报告原始 XML",
            },
            {
                "path": str(dictionary_path),
                "role": "schema_dictionary",
                "description": "字段字典和逻辑表说明",
            },
        ],
        "tables": {
            "PA01": build_pa01(root, xml_path, field_labels),
            "PB01": build_pb01(root, xml_path, field_labels),
            "PB02": build_pb02(root, xml_path, field_labels),
            "PB03": build_pb03(root, xml_path, field_labels),
            "PB04": build_pb04(root, xml_path, field_labels),
            "PC05": build_pc05(root, xml_path, field_labels),
            "PD01": build_pd01(root, xml_path, field_labels),
            "PD03": build_pd03(root, xml_path, field_labels),
            "PH01": build_ph01(root, xml_path, field_labels),
            "PCE": build_pce(root, xml_path, field_labels),
            "POS": build_pos(root, xml_path, field_labels),
        },
        "raw_sections": {
            "source_summary": {
                "pd01_count": len(root.findall(".//PD01")),
                "ph01_count": len(root.findall(".//PH01")),
                "pg01_count": len(root.findall(".//PG01")),
            }
        },
    }
    return document


def validate(instance: dict[str, Any], schema_path: Path) -> None:
    import jsonschema

    schema = json.loads(schema_path.read_text(encoding="utf-8"))
    jsonschema.validate(instance=instance, schema=schema)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xml_path", type=Path, help="Path to individual.xml")
    parser.add_argument("-o", "--output", type=Path, help="Where to write the standard JSON")
    parser.add_argument(
        "--schema",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "schema" / "individual_standard_report.schema.json",
        help="JSON schema used for validation",
    )
    parser.add_argument(
        "--no-validate",
        action="store_true",
        help="Skip schema validation",
    )
    args = parser.parse_args()

    document = build_document(args.xml_path.resolve())
    if not args.no_validate:
        validate(document, args.schema.resolve())

    text = json.dumps(document, ensure_ascii=False, indent=2)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(text + "\n", encoding="utf-8")
    else:
        print(text)


if __name__ == "__main__":
    main()
