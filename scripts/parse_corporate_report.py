#!/usr/bin/env python3
"""Parse corporate credit-report XML into the table-oriented standard JSON schema."""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any
import xml.etree.ElementTree as ET


SCHEMA_VERSION = "0.2.0"
DICTIONARY_FILE = "征信平台信用报告表结构明细V0.3.xlsx"
TABLE_LABEL_SOURCE = DICTIONARY_FILE

TABLE_NAMES = {
    "EA01": "报告头信息单元",
    "EB01": "信贷交易概要信息单元",
    "EC01": "企业基本概况信息单元",
    "EC02": "主要出资人信息单元",
    "EC03": "主要组成人员信息单元",
    "EC04": "上级机构信息单元",
    "EC05": "实际控制人信息单元",
    "ED01": "借贷账户信息单元",
    "ED04": "担保交易信息单元",
    "ED06": "授信协议信息单元",
    "EH01": "评级信息单元",
    "EI01": "异议信息单元",
}

FALLBACK_FIELD_LABELS = {
    "EA01CH": "企业身份标识明细",
    "EA01AI01": "报告编号",
    "EA01AR01": "报告生成时间",
    "EA01BI01": "查询机构代码",
    "EA01BD02": "查询原因",
    "EA01CQ01": "企业名称",
    "EA01CS01": "企业身份标识个数",
    "EA01CD01": "身份标识类型",
    "EA01CI01": "身份标识号码",
    "EA01DS01": "异议标注数目",
    "EC020H": "出资人明细",
    "EC030H": "主要组成人员明细",
    "EC050H": "实际控制人明细",
}

CODED_FIELDS = {
    "EA01BD02",
    "EA01CD01",
    "EC010D01",
    "EC010D02",
    "EC010D03",
    "EC010D04",
    "EC010D05",
    "EC020D01",
    "EC020D03",
    "EC030D01",
    "EC030D02",
    "EC040D01",
    "EC040D02",
    "EC050D01",
    "EC050D02",
    "ED01AD01",
    "ED01AD02",
    "ED01AD04",
    "ED01AD05",
    "ED01AD06",
    "ED01AD07",
    "ED01AD08",
    "ED01AD09",
    "ED01AD10",
    "ED01AD11",
    "ED01BD01",
    "ED01BD02",
    "ED04AD01",
    "ED04AD02",
    "ED04AD03",
    "ED04AD04",
    "ED04AD05",
    "ED04AD06",
    "ED04BD01",
    "ED04BD02",
    "ED04BD04",
    "ED05D01",
    "ED05D02",
    "ED05D03",
    "ED05D04",
    "ED060D01",
    "ED060D02",
    "ED060D03",
    "ED060D04",
    "EH010D01",
    "EI010D01",
    "EI010D02",
}

DATE_FIELDS = {
    "EA01AR01",
    "EC010R01",
    "EC010R02",
    "EC020R01",
    "EC030R01",
    "EC040R01",
    "EC050R01",
    "ED01AR01",
    "ED01AR02",
    "ED01AR03",
    "ED01BR01",
    "ED04AR01",
    "ED04AR02",
    "ED060R01",
    "ED060R02",
    "ED060R03",
    "EH010R01",
    "EI010R01",
}

NUMERIC_FIELDS = {
    "EA01CS01",
    "EA01DS01",
    "EB01AS01",
    "EB01AS02",
    "EB01AJ01",
    "EB01AJ02",
    "EB01AJ03",
    "EB01AJ04",
    "EB01AJ05",
    "EB01AJ06",
    "EB01AJ07",
    "EB01BS01",
    "EB01BS02",
    "EB01BS03",
    "EB01BS04",
    "EB01BS05",
    "EC020J01",
    "EC020S01",
    "EC020Q02",
    "EC030S01",
    "EC050S01",
    "ED01AJ01",
    "ED01BS01",
    "ED01BJ01",
    "ED01BJ02",
    "ED04AJ01",
    "ED04AQ01",
    "ED04BJ01",
    "ED04BJ02",
    "ED050J01",
    "ED050J02",
    "ED050J03",
    "ED050J04",
    "ED050J05",
    "ED060J01",
    "ED060J03",
    "ED060J04",
}


def clean(value: str | None) -> str | None:
    if value is None:
        return None
    value = value.strip()
    return value or None


def parse_date_like(value: str | None) -> date | None:
    value = clean(value)
    if value is None:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y"):
        try:
            parsed = datetime.strptime(value, fmt)
        except ValueError:
            continue
        if fmt == "%Y":
            return date(parsed.year, 1, 1)
        if fmt == "%Y-%m":
            return date(parsed.year, parsed.month, 1)
        return parsed.date()
    return None


def iso_date(value: str | None) -> str | None:
    parsed = parse_date_like(value)
    return parsed.isoformat() if parsed else clean(value)


def to_number(value: str | None) -> int | float | None:
    value = clean(value)
    if value is None:
        return None
    try:
        if "." in value:
            return float(value)
        return int(value)
    except ValueError:
        return None


def transform_value(field_code: str, value: str | None) -> Any:
    if field_code in DATE_FIELDS:
        return iso_date(value)
    if field_code in NUMERIC_FIELDS:
        return to_number(value)
    return clean(value)


def load_dictionary_metadata(base_dir: Path) -> dict[str, str]:
    field_labels = dict(FALLBACK_FIELD_LABELS)
    workbook_path = base_dir / "data_recource" / DICTIONARY_FILE
    try:
        from openpyxl import load_workbook
    except ImportError:
        return field_labels
    if not workbook_path.exists():
        return field_labels
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["企业信用报告表结构"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        _, _, field_code, field_label, *_ = row
        if not field_code:
            continue
        field_code = str(field_code).strip()
        if field_label:
            field_labels[field_code] = str(field_label).strip()
    return field_labels


def make_trace(
    xml_path: Path,
    *,
    source_section: str,
    source_table: str,
    field_code: str | None = None,
    source_id: str | None = None,
) -> dict[str, Any]:
    return {
        "source_file": str(xml_path),
        "source_section": source_section,
        "source_table": source_table,
        "source_fields": [field_code] if field_code else [],
        "source_id": source_id,
    }


def plain_field(field_code: str, field_name: str, value: Any, value_meta: dict[str, Any]) -> dict[str, Any]:
    return {
        "field_code": field_code,
        "field_name": field_name,
        "value": value,
        "value_meta": value_meta,
    }


def coded_field(field_code: str, field_name: str, code: str | None, value_meta: dict[str, Any]) -> dict[str, Any]:
    return {
        "field_code": field_code,
        "field_name": field_name,
        "code": clean(code),
        "label": None,
        "label_source": "unknown",
        "interpretation_status": "unresolved",
        "value_meta": value_meta,
    }


def field_from_text(
    field_code: str,
    raw_value: str | None,
    *,
    xml_path: Path,
    source_section: str,
    source_table: str,
    source_id: str | None,
    field_labels: dict[str, str],
) -> dict[str, Any]:
    field_name = field_labels.get(field_code, field_code)
    value_meta = {
        "value_source": "xml_direct",
        "trace": make_trace(
            xml_path,
            source_section=source_section,
            source_table=source_table,
            field_code=field_code,
            source_id=source_id,
        ),
        "extra_fields": {},
    }
    if field_code in CODED_FIELDS:
        return coded_field(field_code, field_name, raw_value, value_meta)
    return plain_field(field_code, field_name, transform_value(field_code, raw_value), value_meta)


def nested_record(record_index: int, fields: dict[str, Any], record_id: str | None = None) -> dict[str, Any]:
    return {
        "record_index": record_index,
        "record_id": record_id,
        "fields": fields,
    }


def logical_table_single(table_code: str, fields: dict[str, Any]) -> dict[str, Any]:
    return {
        "table_code": table_code,
        "table_name": TABLE_NAMES[table_code],
        "table_label_source": TABLE_LABEL_SOURCE,
        "fields": fields,
    }


def logical_table_repeated(
    table_code: str,
    records: list[dict[str, Any]],
    *,
    summary_fields: dict[str, Any] | None = None,
) -> dict[str, Any]:
    table = {
        "table_code": table_code,
        "table_name": TABLE_NAMES[table_code],
        "table_label_source": TABLE_LABEL_SOURCE,
        "records": records,
    }
    if summary_fields:
        table["summary_fields"] = summary_fields
    return table


def leaf_fields(
    node: ET.Element | None,
    *,
    xml_path: Path,
    source_section: str,
    source_table: str,
    source_id: str | None,
    field_labels: dict[str, str],
    skip_tags: set[str] | None = None,
) -> dict[str, Any]:
    fields: dict[str, Any] = {}
    if node is None:
        return fields
    for child in list(node):
        if len(child) != 0:
            continue
        if skip_tags and child.tag in skip_tags:
            continue
        fields[child.tag] = field_from_text(
            child.tag,
            child.text,
            xml_path=xml_path,
            source_section=source_section,
            source_table=source_table,
            source_id=source_id,
            field_labels=field_labels,
        )
    return fields


def repeated_value_field(
    field_code: str,
    field_name: str,
    records: list[dict[str, Any]],
    *,
    xml_path: Path,
    source_section: str,
    source_table: str,
    source_id: str | None,
) -> dict[str, Any]:
    return plain_field(
        field_code,
        field_name,
        records,
        {
            "value_source": "xml_direct",
            "trace": make_trace(
                xml_path,
                source_section=source_section,
                source_table=source_table,
                field_code=field_code,
                source_id=source_id,
            ),
            "extra_fields": {},
        },
    )


def build_ea01(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    ea01 = root.find(".//EA01")
    fields: dict[str, Any] = {}
    if ea01 is None:
        return logical_table_single("EA01", fields)
    fields.update(
        leaf_fields(
            ea01.find("EA01A"),
            xml_path=xml_path,
            source_section="EAA/EA01/EA01A",
            source_table="ECR_EA01",
            source_id=None,
            field_labels=field_labels,
        )
    )
    fields.update(
        leaf_fields(
            ea01.find("EA01B"),
            xml_path=xml_path,
            source_section="EAA/EA01/EA01B",
            source_table="ECR_EA01",
            source_id=None,
            field_labels=field_labels,
        )
    )
    ea01c = ea01.find("EA01C")
    if ea01c is not None:
        fields.update(
            leaf_fields(
                ea01c,
                xml_path=xml_path,
                source_section="EAA/EA01/EA01C",
                source_table="ECR_EA01",
                source_id=None,
                field_labels=field_labels,
                skip_tags={"EA01CH"},
            )
        )
        nested: list[dict[str, Any]] = []
        for idx, node in enumerate(ea01c.findall("EA01CH"), start=1):
            record_fields = leaf_fields(
                node,
                xml_path=xml_path,
                source_section="EAA/EA01/EA01C/EA01CH",
                source_table="ECR_EA01",
                source_id=f"EA01CH:{idx}",
                field_labels=field_labels,
            )
            nested.append(nested_record(idx, record_fields))
        if nested:
            fields["EA01CH"] = repeated_value_field(
                "EA01CH",
                field_labels.get("EA01CH", "企业身份标识明细"),
                nested,
                xml_path=xml_path,
                source_section="EAA/EA01/EA01C",
                source_table="ECR_EA01",
                source_id=None,
            )
    fields.update(
        leaf_fields(
            ea01.find("EA01D"),
            xml_path=xml_path,
            source_section="EAA/EA01/EA01D",
            source_table="ECR_EA01",
            source_id=None,
            field_labels=field_labels,
        )
    )
    return logical_table_single("EA01", fields)


def build_eb01(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    eb01 = root.find(".//EB01")
    fields: dict[str, Any] = {}
    if eb01 is None:
        return logical_table_single("EB01", fields)
    fields.update(
        leaf_fields(
            eb01.find("EB01A"),
            xml_path=xml_path,
            source_section="EBA/EB01/EB01A",
            source_table="ECR_EB01A",
            source_id=None,
            field_labels=field_labels,
        )
    )
    fields.update(
        leaf_fields(
            eb01.find("EB01B"),
            xml_path=xml_path,
            source_section="EBA/EB01/EB01B",
            source_table="ECR_EB01B",
            source_id=None,
            field_labels=field_labels,
        )
    )
    return logical_table_single("EB01", fields)


def build_ec01(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    return logical_table_single(
        "EC01",
        leaf_fields(
            root.find(".//EC01"),
            xml_path=xml_path,
            source_section="ECA/EC01",
            source_table="ECR_EC01",
            source_id=None,
            field_labels=field_labels,
        ),
    )


def build_ec02(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    ec02 = root.find(".//EC02")
    summary_fields: dict[str, Any] = {}
    records: list[dict[str, Any]] = []
    if ec02 is not None:
        summary_fields = leaf_fields(
            ec02,
            xml_path=xml_path,
            source_section="ECA/EC02",
            source_table="ECR_EC02",
            source_id=None,
            field_labels=field_labels,
            skip_tags={"EC020H"},
        )
        for idx, node in enumerate(ec02.findall("EC020H"), start=1):
            fields = leaf_fields(
                node,
                xml_path=xml_path,
                source_section="ECA/EC02/EC020H",
                source_table="ECR_EC02",
                source_id=f"EC020H:{idx}",
                field_labels=field_labels,
            )
            records.append(nested_record(idx, fields, record_id=clean(node.findtext("EC020Q01"))))
    return logical_table_repeated("EC02", records, summary_fields=summary_fields or None)


def build_ec03(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    ec03 = root.find(".//EC03")
    summary_fields: dict[str, Any] = {}
    records: list[dict[str, Any]] = []
    if ec03 is not None:
        summary_fields = leaf_fields(
            ec03,
            xml_path=xml_path,
            source_section="ECA/EC03",
            source_table="ECR_EC03",
            source_id=None,
            field_labels=field_labels,
            skip_tags={"EC030H"},
        )
        for idx, node in enumerate(ec03.findall("EC030H"), start=1):
            fields = leaf_fields(
                node,
                xml_path=xml_path,
                source_section="ECA/EC03/EC030H",
                source_table="ECR_EC03",
                source_id=f"EC030H:{idx}",
                field_labels=field_labels,
            )
            records.append(nested_record(idx, fields, record_id=clean(node.findtext("EC030Q01"))))
    return logical_table_repeated("EC03", records, summary_fields=summary_fields or None)


def build_ec04(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    return logical_table_single(
        "EC04",
        leaf_fields(
            root.find(".//EC04"),
            xml_path=xml_path,
            source_section="ECA/EC04",
            source_table="ECR_EC04",
            source_id=None,
            field_labels=field_labels,
        ),
    )


def build_ec05(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    ec05 = root.find(".//EC05")
    summary_fields: dict[str, Any] = {}
    records: list[dict[str, Any]] = []
    if ec05 is not None:
        summary_fields = leaf_fields(
            ec05,
            xml_path=xml_path,
            source_section="ECA/EC05",
            source_table="ECR_EC05",
            source_id=None,
            field_labels=field_labels,
            skip_tags={"EC050H"},
        )
        for idx, node in enumerate(ec05.findall("EC050H"), start=1):
            fields = leaf_fields(
                node,
                xml_path=xml_path,
                source_section="ECA/EC05/EC050H",
                source_table="ECR_EC05",
                source_id=f"EC050H:{idx}",
                field_labels=field_labels,
            )
            records.append(nested_record(idx, fields, record_id=clean(node.findtext("EC050Q01"))))
    return logical_table_repeated("EC05", records, summary_fields=summary_fields or None)


def build_ed01(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    for idx, node in enumerate(root.findall(".//ED01"), start=1):
        fields: dict[str, Any] = {}
        fields.update(
            leaf_fields(
                node.find("ED01A"),
                xml_path=xml_path,
                source_section="EDA/ED01/ED01A",
                source_table="ECR_ED01A",
                source_id=f"ED01:{idx}",
                field_labels=field_labels,
            )
        )
        fields.update(
            leaf_fields(
                node.find("ED01B"),
                xml_path=xml_path,
                source_section="EDA/ED01/ED01B",
                source_table="ECR_ED01B",
                source_id=f"ED01:{idx}",
                field_labels=field_labels,
            )
        )
        records.append(nested_record(idx, fields, record_id=clean(node.findtext("./ED01A/ED01AI01"))))
    return logical_table_repeated("ED01", records)


def build_ed04(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    for idx, node in enumerate(root.findall(".//ED04"), start=1):
        fields: dict[str, Any] = {}
        fields.update(
            leaf_fields(
                node.find("ED04A"),
                xml_path=xml_path,
                source_section="EDB/ED04/ED04A",
                source_table="ECR_ED04A",
                source_id=f"ED04:{idx}",
                field_labels=field_labels,
            )
        )
        fields.update(
            leaf_fields(
                node.find("ED04B"),
                xml_path=xml_path,
                source_section="EDB/ED04/ED04B",
                source_table="ECR_ED04B",
                source_id=f"ED04:{idx}",
                field_labels=field_labels,
            )
        )
        records.append(nested_record(idx, fields, record_id=clean(node.findtext("./ED04A/ED04AI01"))))
    return logical_table_repeated("ED04", records)


def build_ed06(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    for idx, node in enumerate(root.findall(".//ED06"), start=1):
        fields = leaf_fields(
            node,
            xml_path=xml_path,
            source_section="EDC/ED06",
            source_table="ECR_ED06",
            source_id=f"ED06:{idx}",
            field_labels=field_labels,
        )
        records.append(nested_record(idx, fields, record_id=clean(node.findtext("ED060I01"))))
    return logical_table_repeated("ED06", records)


def build_eh01(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    for idx, node in enumerate(root.findall(".//EH01"), start=1):
        fields = leaf_fields(
            node,
            xml_path=xml_path,
            source_section="EHA/EH01",
            source_table="ECR_EH01",
            source_id=f"EH01:{idx}",
            field_labels=field_labels,
        )
        records.append(nested_record(idx, fields, record_id=clean(node.findtext("EH010I01"))))
    return logical_table_repeated("EH01", records)


def build_ei01(root: ET.Element, xml_path: Path, field_labels: dict[str, str]) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    for idx, node in enumerate(root.findall(".//EI01"), start=1):
        fields = leaf_fields(
            node,
            xml_path=xml_path,
            source_section="EIA/EI01",
            source_table="ECR_EI01",
            source_id=f"EI01:{idx}",
            field_labels=field_labels,
        )
        records.append(nested_record(idx, fields))
    return logical_table_repeated("EI01", records)


def build_document(xml_path: Path) -> dict[str, Any]:
    base_dir = Path(__file__).resolve().parents[1]
    dictionary_path = base_dir / "data_recource" / DICTIONARY_FILE
    field_labels = load_dictionary_metadata(base_dir)
    root = ET.parse(xml_path).getroot()

    return {
        "schema_version": SCHEMA_VERSION,
        "report_type": "corporate",
        "source_files": [
            {
                "path": str(xml_path),
                "role": "xml",
                "description": "企业征信报告原始 XML",
            },
            {
                "path": str(dictionary_path),
                "role": "schema_dictionary",
                "description": "字段字典和逻辑表说明",
            },
        ],
        "tables": {
            "EA01": build_ea01(root, xml_path, field_labels),
            "EB01": build_eb01(root, xml_path, field_labels),
            "EC01": build_ec01(root, xml_path, field_labels),
            "EC02": build_ec02(root, xml_path, field_labels),
            "EC03": build_ec03(root, xml_path, field_labels),
            "EC04": build_ec04(root, xml_path, field_labels),
            "EC05": build_ec05(root, xml_path, field_labels),
            "ED01": build_ed01(root, xml_path, field_labels),
            "ED04": build_ed04(root, xml_path, field_labels),
            "ED06": build_ed06(root, xml_path, field_labels),
            "EH01": build_eh01(root, xml_path, field_labels),
            "EI01": build_ei01(root, xml_path, field_labels),
        },
        "raw_sections": {
            "source_summary": {
                "ed01_count": len(root.findall(".//ED01")),
                "ed04_count": len(root.findall(".//ED04")),
                "ed05_count": len(root.findall(".//ED05")),
                "ed06_count": len(root.findall(".//ED06")),
                "ei01_count": len(root.findall(".//EI01")),
            }
        },
    }


def validate(instance: dict[str, Any], schema_path: Path) -> None:
    import jsonschema

    schema = json.loads(schema_path.read_text(encoding="utf-8"))
    jsonschema.validate(instance=instance, schema=schema)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xml_path", type=Path, help="Path to corporate.xml")
    parser.add_argument("-o", "--output", type=Path, help="Where to write the standard JSON")
    parser.add_argument(
        "--schema",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "schema" / "corporate_standard_report.schema.json",
        help="JSON schema used for validation",
    )
    parser.add_argument(
        "--no-validate",
        action="store_true",
        help="Skip schema validation",
    )
    args = parser.parse_args()

    document = build_document(args.xml_path.resolve())
    if not args.no_validate:
        validate(document, args.schema.resolve())

    text = json.dumps(document, ensure_ascii=False, indent=2)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(text + "\n", encoding="utf-8")
    else:
        print(text)


if __name__ == "__main__":
    main()
